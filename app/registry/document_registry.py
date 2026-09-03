"""Поиск строк оборудования в Реестр_оборудования.xlsx."""

from __future__ import annotations

import re

from openpyxl import load_workbook

from app import config

_rows_cache: list[dict] | None = None

HEADER_MARKER = "Позиция"


def _load_rows() -> list[dict]:
    global _rows_cache
    if _rows_cache is not None:
        return _rows_cache

    if not config.REGISTRY_XLSX_PATH.exists():
        _rows_cache = []
        return _rows_cache

    wb = load_workbook(config.REGISTRY_XLSX_PATH, read_only=True, data_only=True)
    sheet = wb.worksheets[0]
    rows = list(sheet.iter_rows(values_only=True))

    header_idx = next((i for i, r in enumerate(rows) if r and r[0] == HEADER_MARKER), None)
    if header_idx is None:
        _rows_cache = []
        return _rows_cache

    header = [str(c).strip() if c else "" for c in rows[header_idx]]
    parsed = []
    for row in rows[header_idx + 1:]:
        if not any(row):
            continue
        fields = {h: str(v).strip() for h, v in zip(header, row) if h and v is not None}
        if fields:
            parsed.append(fields)

    _rows_cache = parsed
    return parsed


MATCH_KEYS = (
    "Тип, марка, модель, модификация, характеристики",
    "Заводской номер, артикул",
    "Наименование и техническая характеристика",
)


STOP_WORDS = {
    "какие", "какой", "какая", "какое", "есть", "где", "они", "она", "оно", "он",
    "как", "что", "это", "для", "или", "при", "того", "которые", "который",
    "находится", "находятся", "установлен", "установлены", "можно", "нужно",
}


def _significant_words(text: str) -> list[str]:
    words = re.split(r"\W+", text.lower())
    return [w for w in words if len(w) >= 4 and w not in STOP_WORDS]


def _matches_word(word: str, candidates: list[str]) -> bool:
    """Слово из запроса против слов строки реестра.

    Совпадением считаем общий корень: одно слово начинается с другого.
    Обрезка до пяти символов, как было раньше, роднила «светильник»
    с «светодиодный», а «насос» — с «насосной».
    """
    for cand in candidates:
        if word == cand:
            return True
        shorter, longer = (word, cand) if len(word) <= len(cand) else (cand, word)
        if len(shorter) >= 5 and longer.startswith(shorter) and len(longer) - len(shorter) <= 3:
            return True
    return False


def find_matching_rows(query: str, limit: int = 20) -> list[dict]:
    """Строки реестра, релевантные вопросу — для таблицы под ответом.

    Слово засчитывается по совпадению первых 5 символов, чтобы ловить
    словоформы («лампа» / «лампы» / «ламп»). Одного совпадения мало:
    «насос» находится в «щите управления насосами», поэтому от запроса
    из нескольких слов требуем не меньше половины.
    """
    words = _significant_words(query)
    if not words:
        return []

    rows = _load_rows()
    prepared = []
    for row in rows:
        haystack = " ".join(row.get(k, "") for k in MATCH_KEYS).lower()
        prepared.append((row, [w for w in re.split(r"\W+", haystack) if len(w) >= 3]))

    # слово, которого нет во всём реестре, — это марка или лишнее слово
    # вопроса, а не признак отсутствия оборудования
    known = [w for w in words if any(_matches_word(w, hw) for _, hw in prepared)]
    if not known:
        return []

    required = len(known) if len(known) <= 2 else (len(known) + 1) // 2

    scored: list[tuple[int, dict]] = []
    for row, haystack_words in prepared:
        score = sum(1 for w in known if _matches_word(w, haystack_words))
        if score >= required:
            scored.append((score, row))

    scored.sort(key=lambda x: x[0], reverse=True)
    return [row for _, row in scored[:limit]]
